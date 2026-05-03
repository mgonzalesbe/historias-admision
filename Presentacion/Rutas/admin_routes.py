import re
import time
from io import BytesIO
from flask import Blueprint, render_template, request, session, redirect, url_for, flash, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from Aplicacion.Servicios.HistoriaService import HistoriaService
from Aplicacion.Servicios.AuthService import AuthService
from Aplicacion.Servicios.TokenService import TokenService
from Aplicacion.Servicios.NotificationService import NotificationService, NotificationConfigError
from Aplicacion.Servicios.SolanaBlockchainService import (
    SolanaBlockchainService,
    BlockchainConfigError,
    BlockchainWriteError,
)
from Persistencia.Repositorios.EspecialidadRepository import EspecialidadRepository
from Persistencia.Repositorios.MedicoRepository import MedicoRepository
from Persistencia.Repositorios.PasswordMetricRepository import PasswordMetricRepository
from Persistencia.Repositorios.UsuarioRepository import UsuarioRepository
from Persistencia.Repositorios.HistoriaRepository import HistoriaRepository
from Dominio.Entidades.Especialidad import Especialidad
from Dominio.Entidades.Medico import Medico
from Dominio.Entidades.Usuario import Usuario
from datetime import datetime
import logging
import pyodbc

_logger = logging.getLogger(__name__)


def _build_password_metrics_xlsx(per_user, summary):
    """Genera un .xlsx con la tabla por usuario y un bloque de resumen global."""
    thin = Side(style="thin", color="B4B8C4")
    border_grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    zebra_fill = PatternFill("solid", fgColor="F0F4FA")
    summary_title_fill = PatternFill("solid", fgColor="2563EB")
    summary_title_font = Font(bold=True, color="FFFFFF", size=11)
    summary_label_fill = PatternFill("solid", fgColor="E8EEF7")
    summary_value_fill = PatternFill("solid", fgColor="FFFFFF")

    wb = Workbook()
    ws = wb.active
    ws.title = "Métricas por usuario"
    headers = [
        "Usuario",
        "Nombre",
        "Contraseñas generadas",
        "Caracteres (última contraseña)",
        "Tiempo de generación último (s)",
        "Último registro",
    ]
    ws.append(headers)
    n_users = len(per_user)
    for item in per_user:
        last_at = item.get("last_metric_at")
        last_at_str = last_at.strftime("%d/%m/%Y %H:%M") if last_at else "-"
        ws.append(
            [
                item.get("username") or "",
                item.get("nombre_completo") or "",
                int(item.get("passwords_generated") or 0),
                int(item.get("last_password_length") or 0),
                round((int(item.get("last_generation_ms") or 0)) / 1000.0, 2),
                last_at_str,
            ]
        )
    ws.append([])
    ws.append(["Resumen"])
    total_pw = int(summary.get("total_passwords") or 0)
    avg_len = float(summary.get("avg_length") or 0.0)
    avg_len_rounded = int(avg_len + 0.5)
    avg_sec = round(float(summary.get("avg_generation_ms") or 0.0) / 1000.0, 1)
    ws.append(["Total contraseñas generadas", total_pw])
    ws.append(["Promedio de caracteres (todas las generaciones)", avg_len_rounded])
    ws.append(["Promedio tiempo de generación (s)", avg_sec])

    last_table_row = 1 + n_users
    summary_title_row = 3 + n_users
    summary_last_row = summary_title_row + 3

    # Cabecera tabla
    for col in range(1, 7):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = align_center if col >= 3 else align_left
        c.border = border_grid
    ws.row_dimensions[1].height = 22

    # Filas de datos (cebra)
    for r in range(2, last_table_row + 1):
        zebra = (r % 2 == 0)
        for col in range(1, 7):
            c = ws.cell(row=r, column=col)
            c.border = border_grid
            if zebra:
                c.fill = zebra_fill
            if col in (3, 4, 5):
                c.alignment = align_right
            else:
                c.alignment = align_left

    # Fila en blanco entre tabla y resumen: sin borde obligatorio
    blank_row = last_table_row + 1
    ws.row_dimensions[blank_row].height = 6

    # Bloque resumen
    for col in range(1, 3):
        c = ws.cell(row=summary_title_row, column=col)
        c.font = summary_title_font
        c.fill = summary_title_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_grid
    ws.merge_cells(f"A{summary_title_row}:B{summary_title_row}")
    ws.cell(row=summary_title_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[summary_title_row].height = 20

    for r in range(summary_title_row + 1, summary_last_row + 1):
        for col in range(1, 3):
            c = ws.cell(row=r, column=col)
            c.border = border_grid
            c.alignment = align_left if col == 1 else align_right
            c.fill = summary_label_fill if col == 1 else summary_value_fill
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A2"
    for col_idx in range(1, 7):
        letter = get_column_letter(col_idx)
        scan_end = summary_last_row if col_idx <= 2 else last_table_row
        max_len = 0
        for row_idx in range(1, scan_end + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2.2, 10), 52)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


admin_bp = Blueprint('admin', __name__, url_prefix='/admin')
historia_service = HistoriaService()
auth_service = AuthService()
especialidad_repo = EspecialidadRepository()
medico_repo = MedicoRepository()
usuario_repo = UsuarioRepository()
historia_repo = HistoriaRepository()
password_metric_repo = PasswordMetricRepository()
notification_service = NotificationService()


def _email_is_valid(email: str) -> bool:
    return bool(
        re.fullmatch(
            r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}",
            email or "",
        )
    )


def _password_policy_error(password: str):
    if len(password) < 8:
        return "La contraseña debe tener mínimo 8 caracteres."
    if not re.search(r"[A-Z]", password):
        return "La contraseña debe incluir al menos una mayúscula."
    if not re.search(r"[a-z]", password):
        return "La contraseña debe incluir al menos una minúscula."
    if not re.search(r"[0-9]", password):
        return "La contraseña debe incluir al menos un número."
    if not re.search(r"[^A-Za-z0-9]", password):
        return "La contraseña debe incluir al menos un carácter especial."
    return None


def _password_strength_label(password: str) -> str:
    if _password_policy_error(password) is None:
        return "Fuerte"

    has_length = len(password) >= 8
    has_upper = bool(re.search(r"[A-Z]", password))
    has_lower = bool(re.search(r"[a-z]", password))
    has_digit = bool(re.search(r"[0-9]", password))
    has_special = bool(re.search(r"[^A-Za-z0-9]", password))
    score = sum([has_length, has_upper, has_lower, has_digit, has_special])
    if len(password) >= 12:
        score += 1
    if score >= 5:
        return "Fuerte"
    if score >= 3:
        return "Regular"
    return "Fragil"

@admin_bp.before_request
def check_admin():
    if session.get('usuario_role') != 'admin':
        from flask import redirect, url_for
        return redirect(url_for('auth.login_admin'))

@admin_bp.route('/dashboard')
def dashboard():
    historias = historia_service.get_all_historias()
    cantidades = {
        'total_historias': len(historias),
        'pendientes': len([h for h in historias if h['estado'] == 'Pendiente']),
        'recibidas': len([h for h in historias if h['estado'] == 'Recibido'])
    }
    return render_template('admin/dashboard.html',
                           usuario_nombre=session.get('usuario_nombre'),
                           cantidades=cantidades)

@admin_bp.route('/generar-link', methods=['POST'])
def generar_link():
    token_service = TokenService()
    hours = int(request.form.get('hours', 24))
    token = token_service.generate_token(hours_valid=hours)
    link = request.host_url + 'registro/' + token
    return render_template('admin/generar_link.html', link=link, hours=hours)

@admin_bp.route('/usuarios', methods=['GET', 'POST'])
def usuarios():
    if request.method == 'POST':
        start_ts = time.perf_counter()
        username = (request.form.get('username') or '').strip()
        password = (request.form.get('password') or '').strip()
        nombre_completo = (request.form.get('nombre_completo') or '').strip()
        email = (request.form.get('email') or '').strip()
        role = (request.form.get('role') or '').strip()
        generation_time_ms_raw = (request.form.get('generation_time_ms') or '0').strip()

        try:
            generation_time_ms = max(int(generation_time_ms_raw), 0)
        except ValueError:
            generation_time_ms = 0

        if not username or not password or not nombre_completo or not email:
            flash('Completa todos los campos obligatorios, incluido el correo.', 'danger')
            return redirect(url_for('admin.usuarios'))
        if role not in ('admin', 'admission'):
            flash('Rol inválido.', 'danger')
            return redirect(url_for('admin.usuarios'))
        if not _email_is_valid(email):
            flash('Debes ingresar un correo electrónico válido.', 'danger')
            return redirect(url_for('admin.usuarios'))
        if usuario_repo.get_by_email(email):
            flash('El correo electrónico ya está registrado.', 'warning')
            return redirect(url_for('admin.usuarios'))
        policy_error = _password_policy_error(password)
        if policy_error:
            flash(policy_error, 'danger')
            return redirect(url_for('admin.usuarios'))
        if usuario_repo.get_by_username(username):
            flash('El nombre de usuario ya existe.', 'warning')
            return redirect(url_for('admin.usuarios'))

        try:
            user_id = auth_service.create_user(
                username=username,
                password=password,
                nombre_completo=nombre_completo,
                role=role,
                email=email,
            )
        except pyodbc.Error:
            flash('No se pudo crear el usuario en la base de datos.', 'danger')
            return redirect(url_for('admin.usuarios'))
        except Exception as exc:
            flash(f'Ocurrió un error al crear el usuario: {exc}', 'danger')
            return redirect(url_for('admin.usuarios'))

        print(
            f"[INFO] admin.usuarios created_db user_id={user_id} username={username}",
            flush=True,
        )

        mail_domain = email.split("@")[-1].lower() if "@" in email else ""
        _logger.info(
            "admin.usuarios: usuario creado en BD user_id=%s username=%s role=%s mail_domain=%s",
            user_id,
            username,
            role,
            mail_domain,
        )

        warnings = []

        try:
            salt_hex, commitment = auth_service.build_password_commitment(username, password)
            blockchain_service = SolanaBlockchainService()
            blockchain_service.registrar_password_hash(
                username=username,
                password_commitment=commitment,
                salt_hex=salt_hex,
                role=role,
            )
        except (BlockchainConfigError, BlockchainWriteError) as exc:
            _logger.warning("admin.usuarios: Solana omitido user_id=%s: %s", user_id, exc)
            warnings.append(f'No se pudo registrar el hash en Solana: {exc}')
        except Exception as exc:
            _logger.exception("admin.usuarios: error Solana user_id=%s", user_id)
            warnings.append(f'Error inesperado al registrar hash en Solana: {exc}')

        try:
            password_metric_repo.create(
                usuario_id=user_id,
                password_length=len(password),
                generation_time_ms=generation_time_ms,
                strength_label=_password_strength_label(password),
            )
        except Exception as exc:
            _logger.exception("admin.usuarios: métricas contraseña user_id=%s", user_id)
            warnings.append(f'No se pudieron guardar métricas de contraseña: {exc}')

        try:
            notification_service.send_email_credentials(email, username, password)
        except NotificationConfigError as exc:
            _logger.warning(
                "admin.usuarios: correo credenciales config user_id=%s: %s",
                user_id,
                exc,
            )
            # No bloquear ni mostrar advertencia al usuario final por fallas de correo.
        except Exception as exc:
            _logger.exception(
                "admin.usuarios: correo credenciales falló user_id=%s errno=%s",
                user_id,
                getattr(exc, "errno", None),
            )
            # No bloquear ni mostrar advertencia al usuario final por fallas de correo.

        elapsed_ms = int((time.perf_counter() - start_ts) * 1000)
        print(
            f"[INFO] admin.usuarios finished user_id={user_id} elapsed_ms={elapsed_ms} warnings={len(warnings)}",
            flush=True,
        )

        if warnings:
            flash('Usuario creado con advertencias.', 'warning')
            for warning in warnings:
                flash(warning, 'warning')
        else:
            flash('Usuario creado correctamente. Hash registrado en Solana y credenciales enviadas al correo.', 'success')
        return redirect(url_for('admin.usuarios'))

    return render_template('admin/usuarios.html', usuarios=usuario_repo.get_all())

@admin_bp.route('/usuarios/<int:usuario_id>/editar', methods=['POST'])
def editar_usuario(usuario_id: int):
    usuario_actual = usuario_repo.get_by_id(usuario_id)
    if not usuario_actual:
        flash('Usuario no encontrado.', 'danger')
        return redirect(url_for('admin.usuarios'))

    username = (request.form.get('username') or '').strip()
    password = (request.form.get('password') or '').strip()
    nombre_completo = (request.form.get('nombre_completo') or '').strip()
    email = (request.form.get('email') or '').strip()
    role = (request.form.get('role') or '').strip()
    activo_raw = (request.form.get('activo') or '').strip().lower()
    activo = activo_raw in ('1', 'true', 'on', 'si', 'sí')

    if not username or not nombre_completo or not email:
        flash('Completa los campos obligatorios del usuario, incluido el correo.', 'danger')
        return redirect(url_for('admin.usuarios'))
    if role not in ('admin', 'admission'):
        flash('Rol inválido.', 'danger')
        return redirect(url_for('admin.usuarios'))
    if not _email_is_valid(email):
        flash('Debes ingresar un correo electrónico válido.', 'danger')
        return redirect(url_for('admin.usuarios'))
    if activo_raw not in ('1', '0', 'true', 'false', 'on', 'off', 'si', 'sí', 'no'):
        flash('Estado de usuario inválido.', 'danger')
        return redirect(url_for('admin.usuarios'))
    if session.get('usuario_id') == usuario_id and not activo:
        flash('No puedes desactivar tu propio usuario desde esta pantalla.', 'warning')
        return redirect(url_for('admin.usuarios'))

    existing = usuario_repo.get_by_username(username)
    if existing and existing.id != usuario_id:
        flash('El nombre de usuario ya existe.', 'warning')
        return redirect(url_for('admin.usuarios'))
    existing_email = usuario_repo.get_by_email(email)
    if existing_email and existing_email.id != usuario_id:
        flash('El correo electrónico ya está registrado.', 'warning')
        return redirect(url_for('admin.usuarios'))

    try:
        password_hash = usuario_actual.password_hash
        if password:
            policy_error = _password_policy_error(password)
            if policy_error:
                flash(policy_error, 'danger')
                return redirect(url_for('admin.usuarios'))
            password_hash = auth_service.hash_password(password)

        usuario_actualizado = Usuario(
            id=usuario_actual.id,
            username=username,
            password_hash=password_hash,
            nombre_completo=nombre_completo,
            email=email,
            role=role,
            activo=activo,
            fecha_creacion=usuario_actual.fecha_creacion
        )
        usuario_repo.update(usuario_actualizado)
        flash('Usuario actualizado correctamente.', 'success')
    except pyodbc.Error:
        flash('No se pudo actualizar el usuario.', 'danger')

    return redirect(url_for('admin.usuarios'))

@admin_bp.route('/usuarios/<int:usuario_id>/eliminar', methods=['POST'])
def eliminar_usuario(usuario_id: int):
    if session.get('usuario_id') == usuario_id:
        flash('No puedes eliminar tu propio usuario desde esta pantalla.', 'warning')
        return redirect(url_for('admin.usuarios'))

    try:
        if not usuario_repo.get_by_id(usuario_id):
            flash('Usuario no encontrado.', 'danger')
        elif historia_repo.count_by_usuario_registro(usuario_id) > 0:
            flash('No se puede eliminar el usuario porque tiene historias registradas en admisión.', 'warning')
        else:
            usuario_repo.delete(usuario_id)
            flash('Usuario eliminado correctamente.', 'success')
    except pyodbc.Error:
        flash('No se pudo eliminar el usuario.', 'danger')
    return redirect(url_for('admin.usuarios'))

@admin_bp.route('/historias')
def historias():
    registros = historia_service.get_registros_agrupados_por_dia()
    especialidades = especialidad_repo.get_all()
    return render_template(
        'admin/historias.html',
        registros=registros,
        especialidades=especialidades,
    )


@admin_bp.route('/password-metrics')
def password_metrics():
    try:
        summary = password_metric_repo.get_summary()
        per_user = password_metric_repo.get_per_user()
        per_user.sort(key=lambda item: int(item.get("usuario_id") or 0), reverse=True)
        metrics = password_metric_repo.get_all()
        analytics_data = []
        for metric in metrics:
            created_at = metric.get("created_at")
            analytics_data.append(
                {
                    "created_at": created_at.isoformat() if created_at else None,
                    "username": metric.get("username"),
                    "password_length": int(metric.get("password_length") or 0),
                    "generation_time_ms": int(metric.get("generation_time_ms") or 0),
                    "strength_label": metric.get("strength_label") or "Fragil",
                }
            )
    except Exception as exc:
        print(f"[ERROR] No se pudieron cargar métricas de contraseña: {exc}")
        flash('No se pudieron cargar las métricas de contraseña en este momento.', 'warning')
        summary = {"total_passwords": 0, "avg_length": 0.0, "avg_generation_ms": 0.0}
        per_user = []
        metrics = []
        analytics_data = []
    return render_template(
        'admin/password_metrics.html',
        summary=summary,
        per_user=per_user,
        metrics=metrics,
        analytics_data=analytics_data,
    )


@admin_bp.route('/password-metrics/export')
def password_metrics_export():
    try:
        summary = password_metric_repo.get_summary()
        per_user = password_metric_repo.get_per_user()
        per_user.sort(key=lambda item: int(item.get("usuario_id") or 0), reverse=True)
        buf = _build_password_metrics_xlsx(per_user, summary)
    except Exception as exc:
        _logger.exception("Exportación métricas contraseña: %s", exc)
        flash("No se pudo generar el archivo Excel. Intenta de nuevo.", "danger")
        return redirect(url_for("admin.password_metrics"))
    filename = f"metricas_contrasena_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@admin_bp.route('/historias_json/<fecha>')
def historias_json_admin(fecha):
    fecha_clean = fecha.replace('-', '/')
    try:
        fecha_dt = datetime.strptime(fecha_clean, '%d/%m/%Y').date()
        historias = historia_service.get_historias_por_fecha(fecha_dt)
        return jsonify({'status': 'success', 'historias': historias})
    except Exception as e:
        print(f'Error parsing date {fecha}: {e}')
        return jsonify({'status': 'error', 'historias': []}), 400

@admin_bp.route('/especialidades', methods=['GET', 'POST'])
def especialidades():
    if request.method == 'POST':
        nombre = (request.form.get('nombre') or '').strip()
        descripcion = (request.form.get('descripcion') or '').strip()

        if not nombre:
            flash('El nombre de la especialidad es obligatorio.', 'danger')
            return redirect(url_for('admin.especialidades'))
        if especialidad_repo.get_by_nombre(nombre):
            flash('Ya existe una especialidad con ese nombre.', 'warning')
            return redirect(url_for('admin.especialidades'))

        try:
            especialidad_repo.create(Especialidad.create(nombre, descripcion))
            flash('Especialidad creada correctamente.', 'success')
        except pyodbc.Error:
            flash('No se pudo crear la especialidad.', 'danger')
        return redirect(url_for('admin.especialidades'))

    return render_template('admin/especialidades.html', especialidades=especialidad_repo.get_all())

@admin_bp.route('/especialidades/<int:especialidad_id>/editar', methods=['POST'])
def editar_especialidad(especialidad_id: int):
    especialidad_actual = especialidad_repo.get_by_id(especialidad_id)
    if not especialidad_actual:
        flash('Especialidad no encontrada.', 'danger')
        return redirect(url_for('admin.especialidades'))

    nombre = (request.form.get('nombre') or '').strip()
    descripcion = (request.form.get('descripcion') or '').strip()

    if not nombre:
        flash('El nombre de la especialidad es obligatorio.', 'danger')
        return redirect(url_for('admin.especialidades'))

    existing = especialidad_repo.get_by_nombre(nombre)
    if existing and existing.id != especialidad_id:
        flash('Ya existe una especialidad con ese nombre.', 'warning')
        return redirect(url_for('admin.especialidades'))

    try:
        especialidad_repo.update(Especialidad(id=especialidad_id, nombre=nombre, descripcion=descripcion))
        flash('Especialidad actualizada correctamente.', 'success')
    except pyodbc.Error:
        flash('No se pudo actualizar la especialidad.', 'danger')
    return redirect(url_for('admin.especialidades'))

@admin_bp.route('/especialidades/<int:especialidad_id>/eliminar', methods=['POST'])
def eliminar_especialidad(especialidad_id: int):
    try:
        if not especialidad_repo.get_by_id(especialidad_id):
            flash('Especialidad no encontrada.', 'danger')
        else:
            especialidad_repo.delete(especialidad_id)
            flash('Especialidad eliminada correctamente.', 'success')
    except pyodbc.Error:
        flash('No se puede eliminar la especialidad porque está en uso.', 'warning')
    return redirect(url_for('admin.especialidades'))

@admin_bp.route('/medicos', methods=['GET', 'POST'])
def medicos():
    if request.method == 'POST':
        nombre = (request.form.get('nombre') or '').strip()
        especialidad_id_raw = (request.form.get('especialidad_id') or '').strip()

        if not nombre or not especialidad_id_raw:
            flash('Completa todos los campos del médico.', 'danger')
            return redirect(url_for('admin.medicos'))

        try:
            especialidad_id = int(especialidad_id_raw)
        except ValueError:
            flash('Especialidad inválida.', 'danger')
            return redirect(url_for('admin.medicos'))

        if not especialidad_repo.get_by_id(especialidad_id):
            flash('La especialidad seleccionada no existe.', 'danger')
            return redirect(url_for('admin.medicos'))

        try:
            medico_repo.create(Medico.create(nombre, especialidad_id))
            flash('Médico creado correctamente.', 'success')
        except pyodbc.Error:
            flash('No se pudo crear el médico.', 'danger')
        return redirect(url_for('admin.medicos'))

    especialidades = especialidad_repo.get_all()
    especialidades_map = {esp.id: esp.nombre for esp in especialidades}
    return render_template(
        'admin/medicos.html',
        medicos=medico_repo.get_all(),
        especialidades=especialidades,
        especialidades_map=especialidades_map
    )

@admin_bp.route('/medicos/<int:medico_id>/editar', methods=['POST'])
def editar_medico(medico_id: int):
    medico_actual = medico_repo.get_by_id(medico_id)
    if not medico_actual:
        flash('Médico no encontrado.', 'danger')
        return redirect(url_for('admin.medicos'))

    nombre = (request.form.get('nombre') or '').strip()
    especialidad_id_raw = (request.form.get('especialidad_id') or '').strip()

    if not nombre or not especialidad_id_raw:
        flash('Completa todos los campos del médico.', 'danger')
        return redirect(url_for('admin.medicos'))

    try:
        especialidad_id = int(especialidad_id_raw)
    except ValueError:
        flash('Especialidad inválida.', 'danger')
        return redirect(url_for('admin.medicos'))

    if not especialidad_repo.get_by_id(especialidad_id):
        flash('La especialidad seleccionada no existe.', 'danger')
        return redirect(url_for('admin.medicos'))

    try:
        medico_repo.update(
            Medico(
                id=medico_id,
                nombre=nombre,
                especialidad_id=especialidad_id
            )
        )
        flash('Médico actualizado correctamente.', 'success')
    except pyodbc.Error:
        flash('No se pudo actualizar el médico.', 'danger')

    return redirect(url_for('admin.medicos'))

@admin_bp.route('/medicos/<int:medico_id>/eliminar', methods=['POST'])
def eliminar_medico(medico_id: int):
    try:
        if not medico_repo.get_by_id(medico_id):
            flash('Médico no encontrado.', 'danger')
        elif historia_repo.count_by_medico(medico_id) > 0:
            flash('No se puede eliminar el médico porque tiene historias registradas en admisión.', 'warning')
        else:
            medico_repo.delete(medico_id)
            flash('Médico eliminado correctamente.', 'success')
    except pyodbc.Error:
        flash('No se puede eliminar el médico porque está en uso.', 'warning')
    return redirect(url_for('admin.medicos'))
