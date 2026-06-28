"""
Separa el DNI embebido en columnas de paciente.xlsx.

- En `direccion` y `nombre`: extrae el número a la columna `DNI` y elimina el marcador.
- En `nombrePadre` y `nombreMadre`: elimina marcadores DNI sin moverlos a la columna `DNI`.
- Soporta formatos pegados (12DNI, DNI73858624), `_DNI`, DNI en medio del texto, etc.

Uso (desde la raíz del proyecto):
  python scripts/separar_dni_direccion.py
  python scripts/separar_dni_direccion.py --in-place
  python scripts/separar_dni_direccion.py --input paciente.xlsx --output paciente_limpio.xlsx
"""

from __future__ import annotations

import argparse
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]

DIRECCION_COL = "direccion"
NOMBRE_COL = "nombre"
DNI_COL = "DNI"
PADRE_COL = "nombrePadre"
MADRE_COL = "nombreMadre"

# Variantes: DNI, _DNI, DNI_, DNIN, DNI-19685, DNI, 91512120, DNI:TRAMITE, etc.
DNI_MARKER_WITH_NUMBER = re.compile(
    r"(?:_DNI|DNI_?|DNI[A-Za-z]?)(?:\s*[:;\.,_\-]+)*\s*[Vv]?([\d.]+)",
    re.IGNORECASE,
)
DNI_MARKER_JUNK = re.compile(
    r"(?<![A-Za-z])(?:_DNI|DNI_?|DNI[A-Za-z]?)"
    r"(?:(?:\s*[:;\.,_\-]+)+(?:[A-Za-z]+(?:\s+[A-Za-z]+)*)?|\s+[A-Za-z]+(?:\s+[A-Za-z]+)*)",
    re.IGNORECASE,
)
DNI_MARKER_TAIL = re.compile(r"(?:_DNI|DNI_?|DNI[A-Za-z]?)\s*$", re.IGNORECASE)
DNI_MARKER_GLUED = re.compile(r"DNI(?=\s)", re.IGNORECASE)
DNI_PREFIX_JUNK = re.compile(
    r"^(?:_DNI|DNI_?|DNI[A-Za-z]?)(?:\s*[:;\.,_\-]+)+\s*",
    re.IGNORECASE,
)
DNI_MARKER_ORPHANS = (
    DNI_MARKER_JUNK,
    DNI_MARKER_TAIL,
    DNI_MARKER_GLUED,
)


def _normalize_spaces(texto: str) -> str:
    return re.sub(r"\s+", " ", texto).strip()


def _normalize_dni_number(raw: str) -> str:
    return re.sub(r"\D", "", raw)


TRAILING_DNI = re.compile(r"\s(\d{8})\s*$")


def _extract_dni_number(texto: str) -> str | None:
    sin_prefijo = DNI_PREFIX_JUNK.sub("", texto, count=1)
    prefijo_removido = sin_prefijo != texto

    matches = list(DNI_MARKER_WITH_NUMBER.finditer(sin_prefijo))
    if matches:
        return _normalize_dni_number(matches[-1].group(1)) or None

    if prefijo_removido:
        cola = TRAILING_DNI.search(sin_prefijo)
        if cola:
            return cola.group(1)
    return None


def _remove_dni_markers(texto: str) -> str:
    limpio = DNI_PREFIX_JUNK.sub("", texto, count=1)
    prefijo_removido = limpio != texto
    limpio = DNI_MARKER_WITH_NUMBER.sub("", limpio)
    for patron in DNI_MARKER_ORPHANS:
        limpio = patron.sub("", limpio)
    if prefijo_removido:
        limpio = TRAILING_DNI.sub("", limpio)
    return limpio


def _clean_dni_text(
    valor: str | None, *, extract_number: bool = False
) -> tuple[str | None, str | None]:
    if valor is None:
        return None, None

    texto = str(valor).strip()
    if not texto:
        return texto, None

    dni = _extract_dni_number(texto) if extract_number else None
    limpio = _remove_dni_markers(texto)

    if limpio != texto:
        limpio = _normalize_spaces(limpio)
        return limpio or None, dni

    return texto, dni


def _read_sheet(path: Path) -> tuple[list[str], list[tuple]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h) if h is not None else "" for h in next(rows_iter)]
    data = list(rows_iter)
    wb.close()
    return headers, data


def _write_sheet(path: Path, headers: list[str], data: list[tuple]) -> None:
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("pacientes")
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    ws.append(headers)
    for row in data:
        ws.append(row)
    wb.save(path)
    wb.close()


def procesar_archivo(input_path: Path, output_path: Path) -> dict[str, int]:
    headers, data = _read_sheet(input_path)

    try:
        dir_idx = headers.index(DIRECCION_COL)
    except ValueError as exc:
        raise SystemExit(
            f"No se encontró la columna '{DIRECCION_COL}'. Columnas: {headers}"
        ) from exc

    nombre_idx = headers.index(NOMBRE_COL) if NOMBRE_COL in headers else None
    padre_idx = headers.index(PADRE_COL) if PADRE_COL in headers else None
    madre_idx = headers.index(MADRE_COL) if MADRE_COL in headers else None

    if DNI_COL in headers:
        dni_idx = headers.index(DNI_COL)
        new_headers = headers
    else:
        dni_idx = dir_idx + 1
        new_headers = headers[:dni_idx] + [DNI_COL] + headers[dni_idx:]

    stats = {
        "filas": 0,
        "con_dni": 0,
        "dni_desde_direccion": 0,
        "dni_desde_nombre": 0,
        "sin_dni": 0,
        "nombre_limpiados": 0,
        "padre_limpiados": 0,
        "madre_limpiados": 0,
    }
    new_data: list[tuple] = []

    for row in data:
        row_list = list(row)
        while len(row_list) < len(headers):
            row_list.append(None)

        direccion, dni_dir = _clean_dni_text(row_list[dir_idx], extract_number=True)
        row_list[dir_idx] = direccion

        dni_nom: str | None = None
        if nombre_idx is not None:
            nombre_original = row_list[nombre_idx]
            nombre_limpio, dni_nom = _clean_dni_text(
                row_list[nombre_idx], extract_number=True
            )
            row_list[nombre_idx] = nombre_limpio
            if nombre_limpio != nombre_original:
                stats["nombre_limpiados"] += 1

        if padre_idx is not None:
            padre_limpio, _ = _clean_dni_text(row_list[padre_idx], extract_number=False)
            if padre_limpio != row_list[padre_idx]:
                stats["padre_limpiados"] += 1
            row_list[padre_idx] = padre_limpio

        if madre_idx is not None:
            madre_limpio, _ = _clean_dni_text(row_list[madre_idx], extract_number=False)
            if madre_limpio != row_list[madre_idx]:
                stats["madre_limpiados"] += 1
            row_list[madre_idx] = madre_limpio

        dni = dni_dir or dni_nom
        if dni_dir:
            stats["dni_desde_direccion"] += 1
        elif dni_nom:
            stats["dni_desde_nombre"] += 1

        if DNI_COL in headers:
            row_list[dni_idx] = dni
        else:
            row_list.insert(dni_idx, dni)

        if dni:
            stats["con_dni"] += 1
        stats["filas"] += 1
        new_data.append(tuple(row_list))

    stats["sin_dni"] = stats["filas"] - stats["con_dni"]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    _write_sheet(output_path, new_headers, new_data)

    return stats


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extrae DNI de direccion/nombre y limpia nombrePadre/nombreMadre."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=ROOT / "paciente.xlsx",
        help="Archivo Excel de entrada (default: paciente.xlsx en la raíz).",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Archivo de salida (default: paciente_limpio.xlsx o reemplaza el original con --in-place).",
    )
    parser.add_argument(
        "--in-place",
        action="store_true",
        help="Sobrescribe el archivo de entrada (crea respaldo .bak antes).",
    )
    args = parser.parse_args()

    input_path = args.input.resolve()
    if not input_path.is_file():
        raise SystemExit(f"No existe el archivo: {input_path}")

    if args.in_place:
        backup = input_path.with_suffix(
            f".bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        shutil.copy2(input_path, backup)
        output_path = input_path
        print(f"Respaldo creado: {backup}")
    else:
        output_path = (args.output or ROOT / "paciente_limpio.xlsx").resolve()

    print(f"Leyendo: {input_path}")
    stats = procesar_archivo(input_path, output_path)
    print(f"Guardado: {output_path}")
    print(
        f"Filas procesadas: {stats['filas']} | "
        f"con DNI: {stats['con_dni']} | sin DNI: {stats['sin_dni']}"
    )
    print(
        f"DNI desde direccion: {stats['dni_desde_direccion']} | "
        f"DNI desde nombre: {stats['dni_desde_nombre']}"
    )
    print(
        f"nombre limpiados: {stats['nombre_limpiados']} | "
        f"nombrePadre limpiados: {stats['padre_limpiados']} | "
        f"nombreMadre limpiados: {stats['madre_limpiados']}"
    )


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit(130)
