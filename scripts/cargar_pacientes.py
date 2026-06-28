"""
Carga pacientes desde Excel a Azure SQL / SQL Server.

Uso (desde la raíz del proyecto, con .env apuntando a la base de datos):
  python scripts/cargar_pacientes.py --dry-run
  python scripts/cargar_pacientes.py
  python scripts/cargar_pacientes.py --input paciente_limpio_nuevo.xlsx
  python scripts/cargar_pacientes.py --truncate

Requisito previo: ejecutar config/migration_pacientes.sql en la base de datos.
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    from dotenv import load_dotenv

    load_dotenv(ROOT / ".env")
except ImportError:
    pass

from openpyxl import load_workbook

from Aplicacion.Servicios.PacienteService import PacienteService
from Dominio.Entidades.Paciente import Paciente
from Persistencia.Conexion.DatabaseConnection import DatabaseConnection
from Persistencia.Repositorios.PacienteRepository import PacienteRepository

COL_NOMBRE = "nombre"
COL_FECHA = "fechaNacimiento"
COL_DIRECCION = "direccion"
COL_DNI = "DNI"
COL_PADRE = "nombrePadre"
COL_MADRE = "nombreMadre"
COL_HC = "HistorialClinico"


def _cell_str(valor) -> str | None:
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto or None


def _leer_excel(path: Path) -> list[Paciente]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [
        str(h).strip() if h is not None else ""
        for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    ]
    idx = {name: headers.index(name) for name in [
        COL_NOMBRE, COL_FECHA, COL_DIRECCION, COL_DNI, COL_PADRE, COL_MADRE, COL_HC
    ]}

    pacientes: list[Paciente] = []
    service = PacienteService()

    for row in ws.iter_rows(min_row=2, values_only=True):
        nombre = _cell_str(row[idx[COL_NOMBRE]])
        numero_hc = _cell_str(row[idx[COL_HC]])
        if not nombre or not numero_hc:
            continue

        pacientes.append(
            Paciente.create(
                numero_historia_clinica=numero_hc,
                nombre_completo=nombre,
                dni=_cell_str(row[idx[COL_DNI]]),
                fecha_nacimiento=service._parse_fecha(row[idx[COL_FECHA]]),
                direccion=_cell_str(row[idx[COL_DIRECCION]]),
                nombre_padre=_cell_str(row[idx[COL_PADRE]]),
                nombre_madre=_cell_str(row[idx[COL_MADRE]]),
            )
        )

    wb.close()
    return pacientes


def _truncate_pacientes() -> None:
    db = DatabaseConnection.get_instance()
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM Pacientes")
    conn.commit()
    cursor.execute("DBCC CHECKIDENT ('Pacientes', RESEED, 0)")
    conn.commit()
    cursor.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa pacientes desde Excel a SQL Server.")
    parser.add_argument(
        "--input",
        type=Path,
        default=ROOT / "paciente_limpio_nuevo.xlsx",
        help="Archivo Excel de entrada.",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=500,
        help="Filas por lote al insertar.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Solo lee el Excel y muestra estadísticas, sin insertar.",
    )
    parser.add_argument(
        "--truncate",
        action="store_true",
        help="Vacía la tabla Pacientes antes de importar.",
    )
    args = parser.parse_args()

    input_path = args.input.resolve()
    if not input_path.is_file():
        raise SystemExit(f"No existe el archivo: {input_path}")

    print(f"Leyendo: {input_path}", flush=True)
    pacientes = _leer_excel(input_path)
    print(f"Filas válidas en Excel: {len(pacientes)}", flush=True)

    if args.dry_run:
        con_dni = sum(1 for p in pacientes if p.dni)
        con_padre = sum(1 for p in pacientes if p.nombre_padre)
        con_madre = sum(1 for p in pacientes if p.nombre_madre)
        print(f"Con DNI: {con_dni} | Con padre: {con_padre} | Con madre: {con_madre}")
        print("Dry-run: no se insertó nada.")
        return

    repo = PacienteRepository()
    existentes = repo.count_all()
    if existentes > 0 and not args.truncate:
        print(
            f"La tabla ya tiene {existentes} pacientes. "
            "Use --truncate para vaciarla antes de importar, o importe manualmente."
        )
        raise SystemExit(1)

    if args.truncate and existentes > 0:
        print("Vaciando tabla Pacientes...")
        _truncate_pacientes()

    print("Insertando en la base de datos...", flush=True)
    insertados = repo.bulk_insert(pacientes, batch_size=max(1, args.batch_size))
    total = repo.count_all()
    print(f"Insertados: {insertados} | Total en Pacientes: {total}", flush=True)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit(130)
